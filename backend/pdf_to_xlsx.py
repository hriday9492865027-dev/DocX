import pdfplumber
import openpyxl

def pdf_to_xlsx(pdf_path, xlsx_path):
    wb = openpyxl.Workbook()
    wb.remove(wb.active) # Remove default sheet
    
    with pdfplumber.open(pdf_path) as pdf:
        for i, page in enumerate(pdf.pages):
            sheet = wb.create_sheet(title=f"Page {i+1}")
            tables = page.extract_tables()
            row_idx = 1
            
            for table in tables:
                for row in table:
                    for col_idx, cell in enumerate(row, start=1):
                        sheet.cell(row=row_idx, column=col_idx, value=cell)
                    row_idx += 1
                row_idx += 2 # spacing between tables
                
            if not tables:
                text = page.extract_text()
                if text:
                    for line in text.split('\n'):
                        parts = [p.strip() for p in line.split('   ') if p.strip()]
                        if not parts:
                            parts = line.split('\t')
                        for col_idx, val in enumerate(parts, start=1):
                            sheet.cell(row=row_idx, column=col_idx, value=val)
                        row_idx += 1
                        
    if not wb.sheetnames:
        wb.create_sheet(title="Sheet 1")
    wb.save(xlsx_path)
